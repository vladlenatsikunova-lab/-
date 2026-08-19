"""
Еженедельный отчёт по сторис rog-dog.ru в Telegram — публикуется раз в неделю
(по средам), когда в таблице подходит очередная "Дата публикации".

Источник: Google Sheets "rog-dog_content-plan_v8", лист "Сторис".
Структура листа:
  Неделя | Дата публикации | № слайда | Текст на экране | Стикер | Заметка
Заполнены "Неделя", "Дата публикации" и "Заметка" (тема недели) только в
первой строке каждой недели — у остальных слайдов той же недели эти три
ячейки пустые и относятся к ближайшей заполненной строке сверху.

Нужны переменные окружения:
  ROGDOG_BOT_TOKEN — токен бота
  ROGDOG_CHAT_ID   — id чата (может быть несколько через запятую)
"""

import io
import json
import os
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

SHEET_ID = "1Py-YGK6rzgpd13BR1B1HFSKFT9IAw5T6P_JVRTHp8GA"
SHEET_NAME = "Сторис"
BOT_TOKEN = os.environ["ROGDOG_BOT_TOKEN"]
CHAT_ID = os.environ["ROGDOG_CHAT_ID"]

COL_WEEK = "Неделя"
COL_DATE = "Дата публикации"
COL_SLIDE = "№ слайда"
COL_TEXT = "Текст на экране"
COL_STICKER = "Стикер"
COL_NOTE = "Заметка"

IGNORE_VALUES = {"", "-", "—"}


def download_workbook():
    """Google иногда отвечает на экспорт таблицы дольше обычного —
    пробуем несколько раз с увеличивающимся таймаутом, прежде чем сдаться."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    last_error = None
    for attempt, timeout in enumerate((60, 90, 120), start=1):
        try:
            with urllib.request.urlopen(url, timeout=timeout) as resp:
                data = resp.read()
            return load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            last_error = e
            if attempt < 3:
                time.sleep(10)
    raise last_error


def find_sheet(wb, name: str):
    target = name.strip().lower()
    for sheetname in wb.sheetnames:
        if sheetname.strip().lower() == target:
            return wb[sheetname]
    return None


def parse_row_date(value):
    if isinstance(value, datetime):
        return value.date()
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%d.%m.%Y").date()
    except ValueError:
        return None


def today_msk():
    """Дата, за которую формируем отчёт.

    Если сейчас раньше 6 утра по Москве — считаем, что публикация ещё
    "за вчера" (на случай, если cron-job.org или GitHub Actions запустили отчёт
    с задержкой, уже после полуночи)."""
    now_msk = datetime.now(timezone.utc) + timedelta(hours=3)
    if now_msk.hour < 6:
        now_msk -= timedelta(days=1)
    return now_msk.date()


def build_column_index(ws):
    return {
        (str(cell.value).strip() if cell.value is not None else ""): idx
        for idx, cell in enumerate(ws[1])
    }


def clean(value):
    if value is None:
        return None
    v = str(value).strip()
    if v in IGNORE_VALUES:
        return None
    return v


def collect_weeks(ws, col_index):
    """Проходит по всем строкам листа и группирует слайды по неделям,
    "протягивая" дату/название недели/заметку вниз до следующей заполненной
    строки (стандартный приём для листов с объединёнными по смыслу ячейками)."""
    date_idx = col_index.get(COL_DATE)
    week_idx = col_index.get(COL_WEEK)
    slide_idx = col_index.get(COL_SLIDE)
    text_idx = col_index.get(COL_TEXT)
    sticker_idx = col_index.get(COL_STICKER)
    note_idx = col_index.get(COL_NOTE)

    weeks = []
    current = None

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        raw_date = row[date_idx].value if date_idx is not None and date_idx < len(row) else None
        d = parse_row_date(raw_date)
        if d is not None:
            current = {
                "date": d,
                "week": clean(row[week_idx].value) if week_idx is not None and week_idx < len(row) else None,
                "note": clean(row[note_idx].value) if note_idx is not None and note_idx < len(row) else None,
                "slides": [],
            }
            weeks.append(current)

        if current is None:
            continue

        text = clean(row[text_idx].value) if text_idx is not None and text_idx < len(row) else None
        if not text:
            continue
        sticker = clean(row[sticker_idx].value) if sticker_idx is not None and sticker_idx < len(row) else None
        slide_no = row[slide_idx].value if slide_idx is not None and slide_idx < len(row) else None
        current["slides"].append((slide_no, text, sticker))

    return weeks


def build_message(target_date):
    try:
        wb = download_workbook()
    except Exception as e:
        return f"Не удалось скачать таблицу: {e}"

    ws = find_sheet(wb, SHEET_NAME)
    if ws is None:
        return f"Лист «{SHEET_NAME}» не найден в таблице."

    col_index = build_column_index(ws)
    date_str = target_date.strftime("%d.%m.%Y")
    weeks = collect_weeks(ws, col_index)
    match = next((w for w in weeks if w["date"] == target_date), None)

    if match is None:
        return f"📖 {date_str} — сторис\n\nНа сегодня публикации сторис в таблице не запланировано."

    header = f"📖 {date_str} — сторис"
    if match["week"]:
        header += f" ({match['week']})"
    lines = [header]
    if match["note"]:
        lines.append(match["note"])
    lines.append("")

    for slide_no, text, sticker in match["slides"]:
        prefix = f"{int(slide_no)}. " if isinstance(slide_no, (int, float)) else ""
        line = f"{prefix}{text}"
        if sticker:
            line += f" [{sticker}]"
        lines.append(line)

    return "\n".join(lines).strip()


def send_telegram(text: str):
    # ROGDOG_CHAT_ID может содержать несколько chat_id через запятую —
    # сообщение уходит в каждый из них.
    chat_ids = list(dict.fromkeys(c.strip() for c in CHAT_ID.split(",") if c.strip()))
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    results = []
    for chat_id in chat_ids:
        data = urllib.parse.urlencode({
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": "true",
        }).encode()
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode())
        if not result.get("ok"):
            raise RuntimeError(f"Telegram API error for chat_id={chat_id}: {result}")
        results.append(result)
    return results


if __name__ == "__main__":
    target = today_msk()
    message = build_message(target)
    print(message)
    send_telegram(message)
