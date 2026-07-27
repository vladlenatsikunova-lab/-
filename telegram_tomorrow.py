"""
Дайджест "контент на завтра" в Telegram, со статусом по цвету ячейки:
  зелёная  -> готово к публикации
  жёлтая   -> есть ошибки / на проверке у менеджера
  без цвета -> контента ещё нет / не подтверждено

В отличие от telegram_daily.py (который берёт только текст через CSV-экспорт),
этот скрипт скачивает таблицу как .xlsx (через openpyxl), чтобы прочитать
ещё и заливку (цвет) ячеек — CSV-экспорт цвет не сохраняет.

Нужны переменные окружения:
  TELEGRAM_BOT_TOKEN
  TELEGRAM_CHAT_ID
"""

import io
import json
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

SHEET_ID = "1scpbz_mMTfhSC7q9NDgoNsIi19mRn-GInoDrxm7CLs4"
BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]

MONTHS_RU = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}

SHEETS = ["ЯМАЛМОТО", "Геохакинг", "Дома скучно"]
EMOJI = {"ЯМАЛМОТО": "🏍", "Геохакинг": "🧭", "Дома скучно": "🏠"}

IGNORE_VALUES = {"", "true", "false", "-", "—"}

STATUS_ICON = {"green": "✅", "yellow": "⚠️", "none": "⭕"}


def download_workbook():
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    with urllib.request.urlopen(url, timeout=60) as resp:
        data = resp.read()
    return load_workbook(io.BytesIO(data), data_only=True)


def find_sheet(wb, name: str):
    target = name.strip().lower()
    for sheetname in wb.sheetnames:
        if sheetname.strip().lower() == target:
            return wb[sheetname]
    return None


def parse_ru_date(cell_value):
    """Возвращает (год, месяц, день). Год может быть None, если в ячейке
    просто текст вида '27 июля' без реальной даты (тогда сверяем без года,
    как запасной вариант)."""
    if isinstance(cell_value, datetime):
        return cell_value.year, cell_value.month, cell_value.day
    if not cell_value:
        return None
    m = re.match(r"^\s*(\d{1,2})\s+([а-яёА-ЯЁ]+)", str(cell_value).strip())
    if not m:
        return None
    day = int(m.group(1))
    month = MONTHS_RU.get(m.group(2).lower())
    if month is None:
        return None
    return None, month, day


def date_matches(parsed, target):
    """parsed = (year_or_None, month, day) для ru-листов."""
    if parsed is None:
        return False
    year, month, day = parsed
    if year is None:
        return (month, day) == (target.month, target.day)
    return (year, month, day) == (target.year, target.month, target.day)


def parse_full_date(cell_value):
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if not cell_value:
        return None
    try:
        return datetime.strptime(str(cell_value).strip(), "%d.%m.%Y").date()
    except ValueError:
        return None


def clean(value):
    if value is None:
        return None
    v = str(value).strip()
    if v.lower() in IGNORE_VALUES:
        return None
    return v


def classify_color(cell):
    """Возвращает 'green' / 'yellow' / 'none' по заливке ячейки."""
    fill = cell.fill
    if fill is None or fill.patternType != "solid":
        return "none"
    color = fill.fgColor
    if getattr(color, "type", None) != "rgb" or not color.rgb:
        return "none"
    rgb = color.rgb
    if len(rgb) != 8:
        return "none"
    try:
        r = int(rgb[2:4], 16)
        g = int(rgb[4:6], 16)
        b = int(rgb[6:8], 16)
    except ValueError:
        return "none"
    if g > r + 10 and g > b + 10:
        return "green"
    if r > 200 and g > 200 and b < g - 15 and b < r - 15:
        return "yellow"
    return "none"


def tomorrow_msk():
    today = (datetime.now(timezone.utc) + timedelta(hours=3)).date()
    return today + timedelta(days=1)


def collect_tomorrow_items(ws, sheet_name: str, target_date):
    items = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        date_cell = row[0]
        if sheet_name == "Дома скучно":
            d = parse_full_date(date_cell.value)
            matched = d == target_date
        else:
            parsed = parse_ru_date(date_cell.value)
            matched = date_matches(parsed, target_date)
        if not matched:
            continue
        for cell in row[1:]:
            text = clean(cell.value)
            if not text:
                continue
            status = classify_color(cell)
            items.append((status, text))
    return items


def build_message(target_date):
    lines = [f"🌙 {target_date.strftime('%d.%m.%Y')} — контент на завтра\n"]
    any_content = False
    try:
        wb = download_workbook()
    except Exception as e:
        return f"Не удалось скачать таблицу: {e}"

    for sheet_name in SHEETS:
        ws = find_sheet(wb, sheet_name)
        lines.append(f"{EMOJI[sheet_name]} {sheet_name}:")
        if ws is None:
            lines.append("(лист не найден)")
            lines.append("")
            continue
        items = collect_tomorrow_items(ws, sheet_name, target_date)
        if items:
            any_content = True
            for status, text in items:
                lines.append(f"{STATUS_ICON[status]} {text}")
        else:
            lines.append("нет контента на завтра")
        lines.append("")

    if not any_content:
        lines.append("На завтра по всем аккаунтам пока ничего не готово.")
    lines.append("✅ готово  ⚠️ ошибка/на проверке  ⭕ без статуса")
    return "\n".join(lines).strip()


def send_telegram(text: str):
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    data = urllib.parse.urlencode({
        "chat_id": CHAT_ID,
        "text": text,
        "disable_web_page_preview": "true",
    }).encode()
    req = urllib.request.Request(url, data=data, method="POST")
    with urllib.request.urlopen(req, timeout=30) as resp:
        result = json.loads(resp.read().decode())
    if not result.get("ok"):
        raise RuntimeError(f"Telegram API error: {result}")
    return result


if __name__ == "__main__":
    target = tomorrow_msk()
    message = build_message(target)
    print(message)
    send_telegram(message)
