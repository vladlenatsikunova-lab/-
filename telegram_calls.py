"""
Отчёт по обзвонам клиентской базы за сегодня — в Telegram, каждый день в 22:00 МСК.

Источник: Google Sheets "Скрипт и трекер обзвона | ЯМАЛМОТО", лист "Отчёт по дням".
Структура листа:
  - колонка A — дата (строка данных начинается с 6-й строки)
  - строка 4 — имена менеджеров, объединённые ячейки (у каждого менеджера
    может быть НЕСКОЛЬКО отдельных блоков колонок — старые метрики и новые
    метрики лежат раздельно, но относятся к одному и тому же менеджеру)
  - строка 5 — подписи метрик на каждую колонку
  - в конце листа есть отдельный блок "ИТОГО (весь отдел)" — это НЕ менеджер,
    а сумма по всем менеджерам; скрипт его пропускает и считает итоги сам,
    чтобы не задваивать цифры.
Скрипт сам вычитывает имена менеджеров и колонки из строк 4-5, поэтому если
менеджеров или метрик станет больше/меньше — ничего в коде менять не надо.

Нужны переменные окружения:
  CALLS_BOT_TOKEN — токен бота (@ai_yamalmoto_baza_bot)
  CALLS_CHAT_ID   — id чата (может быть несколько через запятую)
"""

import io
import json
import os
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

SHEET_ID = "1_rYOoPJZfuOh4GHdsL8NBDmWovu9DqsdLSJuU4BxgwE"
SHEET_NAME = "Отчёт по дням"
BOT_TOKEN = os.environ["CALLS_BOT_TOKEN"]
CHAT_ID = os.environ["CALLS_CHAT_ID"]

HEADER_ROW_NAMES = 4   # строка с именами менеджеров (объединённые ячейки)
HEADER_ROW_METRICS = 5  # строка с названиями метрик
DATA_START_ROW = 6

TOTAL_BLOCK_PREFIX = "ИТОГО"  # так начинается заголовок агрегированного блока по отделу

# короткие подписи метрик для сообщения
SHORT_LABELS = {
    "Обзвонено": "Обзвонено",
    "Взяли трубку": "Взяли трубку",
    "Не взяли / не беспокоить": "Не взяли",
    "Негатив": "Негатив",
    "Вступили в закрытый чат": "Вступили в чат",
    "Запросы на технику": "Запросы техника",
    "Запросы на запчасти": "Запросы запчасти",
    "Продажи запчастей": "Продажи запчастей",
    "Продажи техники": "Продажи техники",
}


def fmt_num(value):
    """Google Sheets отдаёт целые числа как float (0.0, 3.0) — приводим к int."""
    if value is None:
        return 0
    if isinstance(value, float) and value.is_integer():
        return int(value)
    return value


def download_workbook():
    """Google иногда отвечает на экспорт таблицы дольше обычного —
    пробуем несколько раз с увеличивающимся таймаутом, прежде чем сдаться."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    last_error = None
    for attempt, timeout in enumerate((60, 90, 120), start=1):
        try:
            with urllib.request.urlopen(url, timeout=timeout) as resp:
                data = resp.read()
            return load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            last_error = e
            if attempt < 3:
                time.sleep(10)
    raise last_error


def find_sheet(wb, name: str):
    target = name.strip().lower()
    for sheetname in wb.sheetnames:
        if sheetname.strip().lower() == target:
            return wb[sheetname]
    return None


def build_column_map(ws):
    """Возвращает список (col_idx, manager_name, metric_label) для всех колонок
    с данными по менеджерам. Колонки агрегированного блока "ИТОГО (весь отдел)"
    сюда не попадают — этот блок пропускается, так как итоги считаются сами
    по факту суммирования данных менеджеров (см. build_message)."""
    columns = []
    current_manager = None
    max_col = ws.max_column
    for col in range(2, max_col + 1):
        name_cell = ws.cell(row=HEADER_ROW_NAMES, column=col).value
        if name_cell:
            current_manager = str(name_cell).strip()
        metric_cell = ws.cell(row=HEADER_ROW_METRICS, column=col).value
        if not metric_cell or current_manager is None:
            continue
        if current_manager.upper().startswith(TOTAL_BLOCK_PREFIX):
            continue  # это блок "ИТОГО (весь отдел)", а не менеджер — пропускаем
        columns.append((col, current_manager, str(metric_cell).strip()))
    return columns


def today_msk():
    """Дата, за которую формируем отчёт.

    GitHub Actions иногда запускает cron с большой задержкой (бывает,
    что вместо 21:00 скрипт стартует после полуночи). Если это произошло,
    "сегодня" по часам уже успело смениться на новый день, за который
    менеджеры физически ещё не могли ничего внести — и отчёт улетал бы
    пустым, с одними нулями. Чтобы так не было: если сейчас раньше 6 утра
    по Москве, считаем, что отчёт всё ещё "за вчера" (просто опоздавший),
    и берём предыдущую календарную дату.
    """
    now_msk = datetime.now(timezone.utc) + timedelta(hours=3)
    if now_msk.hour < 6:
        now_msk -= timedelta(days=1)
    return now_msk.date()


def find_today_row(ws, target_date):
    for row in range(DATA_START_ROW, ws.max_row + 1):
        cell_value = ws.cell(row=row, column=1).value
        d = None
        if isinstance(cell_value, datetime):
            d = cell_value.date()
        elif hasattr(cell_value, "year"):  # date object
            d = cell_value
        if d == target_date:
            return row
    return None


def build_message(target_date):
    try:
        wb = download_workbook()
    except Exception as e:
        return f"Не удалось скачать таблицу: {e}"

    ws = find_sheet(wb, SHEET_NAME)
    if ws is None:
        return f"Лист «{SHEET_NAME}» не найден в таблице."

    columns = build_column_map(ws)
    row = find_today_row(ws, target_date)

    lines = [f"\U0001F4DE Отчёт по обзвонам за {target_date.strftime('%d.%m.%Y')}", ""]

    if row is None:
        lines.append("Строка с сегодняшней датой не найдена в таблице.")
        return "\n".join(lines).strip()

    # Группируем колонки по менеджеру. Один и тот же менеджер может встречаться
    # в НЕСКОЛЬКИХ блоках колонок (старые метрики + новые метрики лежат отдельно) —
    # поэтому группируем по имени, а не по соседству колонок, чтобы все метрики
    # одного менеджера попали в один блок сообщения.
    managers_order = []
    managers_data = {}
    for col, manager, metric in columns:
        if manager not in managers_data:
            managers_data[manager] = []
            managers_order.append(manager)
        value = fmt_num(ws.cell(row=row, column=col).value)
        managers_data[manager].append((metric, value))

    totals = {}
    for manager in managers_order:
        metrics = managers_data[manager]
        lines.append(f"👤 {manager}")
        parts = []
        for metric, value in metrics:
            label = SHORT_LABELS.get(metric, metric)
            parts.append(f"{label}: {value}")
            if isinstance(value, (int, float)):
                totals[label] = totals.get(label, 0) + value
        lines.append("   " + " · ".join(parts))
        lines.append("")

    if totals:
        total_parts = [f"{label}: {fmt_num(value)}" for label, value in totals.items()]
        lines.append("\U0001F4CA Итого по отделу: " + " · ".join(total_parts))

    return "\n".join(lines).strip()


def send_telegram(text: str):
    # dict.fromkeys вместо set — убирает дубликаты id (если в CALLS_CHAT_ID
    # один и тот же чат случайно указан дважды), сохраняя порядок
    chat_ids = list(dict.fromkeys(c.strip() for c in CHAT_ID.split(",") if c.strip()))
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    results = []
    for chat_id in chat_ids:
        data = urllib.parse.urlencode({
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": "true",
        }).encode()
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode())
        if not result.get("ok"):
            raise RuntimeError(f"Telegram API error for chat_id={chat_id}: {result}")
        results.append(result)
    return results


if __name__ == "__main__":
    target = today_msk()
    message = build_message(target)
    print(message)
    send_telegram(message)
