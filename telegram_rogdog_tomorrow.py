"""
Ежедневный отчёт по контент-плану rog-dog.ru (Reels/Stories) в Telegram —
что нужно ПОДГОТОВИТЬ НА ЗАВТРА. Присылается днём, в 15:00 МСК.

Источник: Google Sheets "rog-dog_content-plan_v8", лист "КП-план".
Структура листа (по столбцам):
  № | Дата | День | Reels/Stories | Материал (папка Я.Диска) | Формат |
  Стиль хука | Тема ролика (хук) | Оффер в CTA
Дата в таблице хранится текстом вида "17.08.2026".

Нужны переменные окружения:
  ROGDOG_BOT_TOKEN — токен бота
  ROGDOG_CHAT_ID   — id чата (может быть несколько через запятую)
"""

import io
import json
import os
import time
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

SHEET_ID = "1Py-YGK6rzgpd13BR1B1HFSKFT9IAw5T6P_JVRTHp8GA"
SHEET_NAME = "КП-план"
BOT_TOKEN = os.environ["ROGDOG_BOT_TOKEN"]
CHAT_ID = os.environ["ROGDOG_CHAT_ID"]

COL_DATE = "Дата"
COL_DAY = "День"
COL_TYPE = "Reels/Stories"
COL_MATERIAL = "Материал (папка Я.Диска)"
COL_FORMAT = "Формат"
COL_HOOK_STYLE = "Стиль хука"
COL_TOPIC = "Тема ролика (хук)"
COL_OFFER = "Оффер в CTA"


def download_workbook():
    """Google иногда отвечает на экспорт таблицы дольше обычного —
    пробуем несколько раз с увеличивающимся таймаутом, прежде чем сдаться."""
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    last_error = None
    for attempt, timeout in enumerate((60, 90, 120), start=1):
        try:
            with urllib.request.urlopen(url, timeout=timeout) as resp:
                data = resp.read()
            return load_workbook(io.BytesIO(data), data_only=True)
        except Exception as e:
            last_error = e
            if attempt < 3:
                time.sleep(10)
    raise last_error


def find_sheet(wb, name: str):
    target = name.strip().lower()
    for sheetname in wb.sheetnames:
        if sheetname.strip().lower() == target:
            return wb[sheetname]
    return None


def parse_row_date(value):
    if isinstance(value, datetime):
        return value.date()
    if not value:
        return None
    try:
        return datetime.strptime(str(value).strip(), "%d.%m.%Y").date()
    except ValueError:
        return None


def tomorrow_msk():
    today = (datetime.now(timezone.utc) + timedelta(hours=3)).date()
    return today + timedelta(days=1)


def build_column_index(ws):
    return {
        (str(cell.value).strip() if cell.value is not None else ""): idx
        for idx, cell in enumerate(ws[1])
    }


def find_row_for_date(ws, col_index, target_date):
    date_idx = col_index.get(COL_DATE)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if date_idx is None or date_idx >= len(row):
            continue
        d = parse_row_date(row[date_idx].value)
        if d == target_date:
            return row
    return None


def cell_text(row, col_index, name):
    idx = col_index.get(name)
    if idx is None or idx >= len(row):
        return "—"
    value = row[idx].value
    if value is None or str(value).strip() == "":
        return "—"
    return str(value).strip()


def build_message(target_date, label):
    try:
        wb = download_workbook()
    except Exception as e:
        return f"Не удалось скачать таблицу: {e}"

    ws = find_sheet(wb, SHEET_NAME)
    if ws is None:
        return f"Лист «{SHEET_NAME}» не найден в таблице."

    col_index = build_column_index(ws)
    row = find_row_for_date(ws, col_index, target_date)
    date_str = target_date.strftime("%d.%m.%Y")

    if row is None:
        return f"📅 {date_str} — {label}\n\nСтрока с этой датой не найдена в «{SHEET_NAME}»."

    day = cell_text(row, col_index, COL_DAY)
    content_type = cell_text(row, col_index, COL_TYPE)
    material = cell_text(row, col_index, COL_MATERIAL)
    fmt = cell_text(row, col_index, COL_FORMAT)
    hook_style = cell_text(row, col_index, COL_HOOK_STYLE)
    topic = cell_text(row, col_index, COL_TOPIC)
    offer = cell_text(row, col_index, COL_OFFER)

    lines = [
        f"📅 {date_str} ({day}) — {label}",
        "",
        f"Публикация: {content_type}",
        f"Материал: {material}",
        f"Формат: {fmt}",
        f"Стиль хука: {hook_style}",
        f"Тема ролика: {topic}",
        f"Оффер в CTA: {offer}",
    ]
    return "\n".join(lines).strip()


def send_telegram(text: str):
    # ROGDOG_CHAT_ID может содержать несколько chat_id через запятую —
    # сообщение уходит в каждый из них.
    chat_ids = list(dict.fromkeys(c.strip() for c in CHAT_ID.split(",") if c.strip()))
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    results = []
    for chat_id in chat_ids:
        data = urllib.parse.urlencode({
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": "true",
        }).encode()
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode())
        if not result.get("ok"):
            raise RuntimeError(f"Telegram API error for chat_id={chat_id}: {result}")
        results.append(result)
    return results


if __name__ == "__main__":
    target = tomorrow_msk()
    message = build_message(target, "план на завтра")
    print(message)
    send_telegram(message)
