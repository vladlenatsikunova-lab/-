"""
Ежедневный дайджест контент-плана в Telegram.

Берёт данные из Google Sheets (три листа: ЯМАЛМОТО, Геохакинг, Дома скучно),
находит строки с сегодняшней датой (колонка A, сравнение по ПОЛНОЙ дате
включая год — чтобы не путать одинаковые "27 июля" из разных лет) и
присылает сообщение в Telegram со всем, что запланировано на сегодня.

Таблица читается как .xlsx (через openpyxl), а не CSV — так колонка A
отдаёт настоящие даты (день/месяц/год), а не отформатированный текст без года.

Нужны переменные окружения:
  TELEGRAM_BOT_TOKEN — токен бота от @BotFather
  TELEGRAM_CHAT_ID   — id чата/канала или @username публичного канала

Запуск: python telegram_daily.py
"""

import io
import json
import os
import re
import urllib.parse
import urllib.request
from datetime import datetime, timedelta, timezone

from openpyxl import load_workbook

SHEET_ID = "1scpbz_mMTfhSC7q9NDgoNsIi19mRn-GInoDrxm7CLs4"
BOT_TOKEN = os.environ["TELEGRAM_BOT_TOKEN"]
CHAT_ID = os.environ["TELEGRAM_CHAT_ID"]

MONTHS_RU = {
    "января": 1, "февраля": 2, "марта": 3, "апреля": 4,
    "мая": 5, "июня": 6, "июля": 7, "августа": 8,
    "сентября": 9, "октября": 10, "ноября": 11, "декабря": 12,
}

SHEETS = ["ЯМАЛМОТО", "Геохакинг", "Дома скучно"]
EMOJI = {"ЯМАЛМОТО": "🏍", "Геохакинг": "🧭", "Дома скучно": "🏠"}

# значения, которые не считаются реальным контентом (чекбоксы, прочерки и т.п.)
IGNORE_VALUES = {"", "true", "false", "-", "—"}


def download_workbook():
    url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=xlsx"
    with urllib.request.urlopen(url, timeout=60) as resp:
        data = resp.read()
    return load_workbook(io.BytesIO(data), data_only=True)


def find_sheet(wb, name: str):
    target = name.strip().lower()
    for sheetname in wb.sheetnames:
        if sheetname.strip().lower() == target:
            return wb[sheetname]
    return None


def parse_ru_date(cell_value):
    """Возвращает (год, месяц, день). Год может быть None, если в ячейке
    просто текст вида '27 июля' без реальной даты (тогда сверяем без года,
    как запасной вариант)."""
    if isinstance(cell_value, datetime):
        return cell_value.year, cell_value.month, cell_value.day
    if not cell_value:
        return None
    m = re.match(r"^\s*(\d{1,2})\s+([а-яёА-ЯЁ]+)", str(cell_value).strip())
    if not m:
        return None
    day = int(m.group(1))
    month = MONTHS_RU.get(m.group(2).lower())
    if month is None:
        return None
    return None, month, day


def parse_full_date(cell_value):
    if isinstance(cell_value, datetime):
        return cell_value.date()
    if not cell_value:
        return None
    try:
        return datetime.strptime(str(cell_value).strip(), "%d.%m.%Y").date()
    except ValueError:
        return None


def clean(value):
    if value is None:
        return None
    v = str(value).strip()
    if v.lower() in IGNORE_VALUES:
        return None
    return v


def today_msk():
    return (datetime.now(timezone.utc) + timedelta(hours=3)).date()


def date_matches(parsed, target):
    """parsed = (year_or_None, month, day) для ru-листов."""
    if parsed is None:
        return False
    year, month, day = parsed
    if year is None:
        return (month, day) == (target.month, target.day)
    return (year, month, day) == (target.year, target.month, target.day)


def collect_today_content(ws, sheet_name: str, target_date):
    items = []
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        date_cell = row[0].value
        if sheet_name == "Дома скучно":
            d = parse_full_date(date_cell)
            matched = d == target_date
        else:
            parsed = parse_ru_date(date_cell)
            matched = date_matches(parsed, target_date)
        if not matched:
            continue
        for cell in row[1:]:
            v = clean(cell.value)
            if v and v not in items:
                items.append(v)
    return items


def build_message(target_date):
    lines = [f"\U0001F4C5 {target_date.strftime('%d.%m.%Y')} — контент-план на сегодня\n"]
    any_content = False
    try:
        wb = download_workbook()
    except Exception as e:
        return f"Не удалось скачать таблицу: {e}"

    for sheet_name in SHEETS:
        ws = find_sheet(wb, sheet_name)
        lines.append(f"{EMOJI[sheet_name]} {sheet_name}:")
        if ws is None:
            lines.append("(лист не найден)")
            lines.append("")
            continue
        items = collect_today_content(ws, sheet_name, target_date)
        if items:
            any_content = True
            for it in items:
                lines.append(f"- {it}")
        else:
            lines.append("нет публикаций сегодня")
        lines.append("")
    if not any_content:
        lines.append("На сегодня по всем аккаунтам публикаций не найдено.")
    return "\n".join(lines).strip()


def send_telegram(text: str):
    # TELEGRAM_CHAT_ID может содержать несколько chat_id через запятую —
    # сообщение уходит в каждый из них.
    chat_ids = [c.strip() for c in CHAT_ID.split(",") if c.strip()]
    url = f"https://api.telegram.org/bot{BOT_TOKEN}/sendMessage"
    results = []
    for chat_id in chat_ids:
        data = urllib.parse.urlencode({
            "chat_id": chat_id,
            "text": text,
            "disable_web_page_preview": "true",
        }).encode()
        req = urllib.request.Request(url, data=data, method="POST")
        with urllib.request.urlopen(req, timeout=30) as resp:
            result = json.loads(resp.read().decode())
        if not result.get("ok"):
            raise RuntimeError(f"Telegram API error for chat_id={chat_id}: {result}")
        results.append(result)
    return results


if __name__ == "__main__":
    today = today_msk()
    message = build_message(today)
    print(message)
    send_telegram(message)
